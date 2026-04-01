from __future__ import annotations

from pathlib import Path
from typing import List, Optional, Union
import openpyxl
import logging

from legacy.features import ISI_time, StartEndFreq, StartEndFreq_from_paths


def _find_column(worksheet, name: str) -> Optional[int]:
    """Return the 1-based column index for *name* in row 1, or None."""
    for col in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=col).value == name:
            return col
    return None


def compute_basic_features(
    file_path: str,
    signal_vec: Optional[List],
    siz: int,
    mother: List,
    name: List,
    age: List,
    session: List,
    rec_num: List,
    mother_syl: List,
    name_syl: List,
    age_syl: List,
    session_syl: List,
    rec_num_syl: List,
    start_syl: List,
    end_syl: List,
    rate: int,
    logger: Optional[logging.Logger] = None,
    audio_paths: Optional[List[Union[str, Path]]] = None,
) -> str:
    """
    Compute basic features (ISI time and start/end frequencies) and write to Excel.
    
    This function:
    1. Computes ISI (Inter-Syllable Interval) time
    2. Computes start and end frequencies for each syllable
    3. Writes the results to the Excel file as new columns
    
    Idempotent: if the columns already exist they are overwritten in place.
    
    Args:
        file_path: Path to the segmentation Excel file (will be updated)
        signal_vec: List of audio signal arrays (used if ``audio_paths`` is None)
        siz: Number of recordings
        mother, name, age, session, rec_num: Metadata lists for recordings
        mother_syl, name_syl, age_syl, session_syl, rec_num_syl: Metadata lists for syllables
        start_syl, end_syl: Start and end times for each syllable
        rate: Sampling rate
        logger: Optional logger instance for logging
    
    Returns:
        Path to the updated Excel file
    """
    if logger:
        logger.info("Computing basic features: ISI time and start/end frequencies")
    
    ISI = ISI_time(rec_num_syl, start_syl, end_syl)
    if audio_paths is not None:
        startF, endF = StartEndFreq_from_paths(
            audio_paths,
            siz,
            mother,
            name,
            age,
            session,
            rec_num,
            mother_syl,
            name_syl,
            age_syl,
            session_syl,
            rec_num_syl,
            start_syl,
            end_syl,
            rate,
            logger=logger,
        )
    else:
        if signal_vec is None:
            raise ValueError("compute_basic_features requires signal_vec or audio_paths")
        startF, endF = StartEndFreq(
            signal_vec,
            siz,
            mother,
            name,
            age,
            session,
            rec_num,
            mother_syl,
            name_syl,
            age_syl,
            session_syl,
            rec_num_syl,
            start_syl,
            end_syl,
            rate,
            logger=logger,
        )
    
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.worksheets[0]
    
    column_names = ['ISI_time', 'Start Point (Hz)', 'End Point (Hz)']
    data_lists = [ISI, startF, endF]
    
    next_new = worksheet.max_column + 1
    for col_name, data in zip(column_names, data_lists):
        col = _find_column(worksheet, col_name)
        if col is None:
            col = next_new
            next_new += 1
        worksheet.cell(row=1, column=col).value = col_name
        for row_idx, value in enumerate(data, start=2):
            worksheet.cell(row=row_idx, column=col).value = value
    
    workbook.save(file_path)
    
    if logger:
        logger.info(f"Basic features computed and written to {file_path}")
    
    return file_path
