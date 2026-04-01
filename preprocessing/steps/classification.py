from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Callable, List, Optional, Tuple
import logging
import os

import numpy as np
import openpyxl
from tensorflow import keras

from legacy.statistics_generator import Syl_Class_Vec
from utils import replace_extension

LOW_CONFIDENCE_CLASS = 10
CONFIDENCE_THRESHOLD = 0.5

# Some bundles use variables-001.data-* while TensorFlow expects variables.data-*.
_VAR_SHARD_RE = re.compile(r"^variables-\d+\.(data-\d+-of-\d+)$")
_VAR_INDEX_RE = re.compile(r"^variables-\d+\.index$")


def _ensure_savedmodel_variable_filenames(model_dir: Path) -> None:
    """
    If ``variables/`` contains only ``variables-001.data-...`` shards, TF still opens
    ``variables.data-...`` (per ``variables.index``). Copy missing canonical names.
    """
    var_dir = model_dir / "variables"
    if not var_dir.is_dir():
        return
    for f in list(var_dir.iterdir()):
        if not f.is_file():
            continue
        m = _VAR_SHARD_RE.match(f.name)
        if m:
            target = var_dir / f"variables.{m.group(1)}"
            if not target.exists():
                shutil.copy2(f, target)
            continue
        m = _VAR_INDEX_RE.match(f.name)
        if m:
            target = var_dir / "variables.index"
            if not target.exists():
                shutil.copy2(f, target)


def load_classification_model(model_path: str):
    """Load a pre-trained Keras / TensorFlow SavedModel from a file or directory path."""
    p = Path(model_path).expanduser()
    if p.is_dir() and (p / "saved_model.pb").is_file():
        _ensure_savedmodel_variable_filenames(p)
    return keras.models.load_model(model_path)


def classify_syllables(
    year: str,
    model,
    age_syl: List,
    matgen_syl: List,
    pupgen_syl: List,
    mother_syl: List,
    name_syl: List,
    sex_syl: List,
    session_syl: List,
    rec_num_syl: List,
    start_syl: List,
    end_syl: List,
    logger: Optional[logging.Logger] = None,
    year_audio_root: Optional[Path] = None,
    progress_hook: Optional[Callable[[int, int], None]] = None,
) -> np.ndarray:
    """Classify each syllable by generating spectrograms and running the CNN model.

    For each syllable: loads the WAV, extracts the audio segment, applies a 30kHz
    high-pass filter, computes the STFT spectrogram, resizes to 128x128, and feeds
    it to the model. Results are grouped per recording as `sample` objects.
    """
    return Syl_Class_Vec(
        year,
        model,
        age_syl,
        matgen_syl,
        pupgen_syl,
        mother_syl,
        name_syl,
        sex_syl,
        session_syl,
        rec_num_syl,
        start_syl,
        end_syl,
        logger=logger,
        year_audio_root=year_audio_root,
        progress_hook=progress_hook,
    )


def save_raw_predictions(file_path: str, samples: np.ndarray) -> str:
    """Save the raw sample predictions to a .npy file alongside the Excel output.

    Returns the path to the saved .npy file.
    """
    output_npy = replace_extension(file_path, ".npy")
    np.save(output_npy, samples)
    return output_npy


def postprocess_predictions(
    samples: np.ndarray,
    logger: Optional[logging.Logger] = None,
) -> List[int]:
    """Convert raw model probabilities into syllable type numbers (0-10).

    For each syllable prediction: if the maximum probability is below the
    confidence threshold, assign class 10 (noise/unknown). Otherwise, take
    the argmax (0-9) as the syllable type.
    """
    syl_num: List[int] = []
    for i in range(len(samples)):
        for j in range(len(samples[i].syls)):
            if np.max(samples[i].syls[j]) < CONFIDENCE_THRESHOLD:
                assigned = LOW_CONFIDENCE_CLASS
            else:
                assigned = int(np.argmax(samples[i].syls[j]))
            samples[i].syls[j] = assigned
            syl_num.append(assigned)
            if logger:
                logger.debug(f"Syllable number: {assigned}")
    return syl_num


def _find_column(worksheet, name: str) -> Optional[int]:
    """Return the 1-based column index for *name* in row 1, or None."""
    for col in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=col).value == name:
            return col
    return None


def write_syllable_numbers(file_path: str, syl_num: List[int]) -> None:
    """Write 'Syllable number' column to the segmentation Excel file.

    Idempotent: if the column already exists it is overwritten in place.
    """
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.worksheets[0]
    col = _find_column(worksheet, "Syllable number")
    if col is None:
        col = worksheet.max_column + 1
    worksheet.cell(row=1, column=col).value = "Syllable number"
    for idx, syl_val in enumerate(syl_num, start=2):
        worksheet.cell(row=idx, column=col).value = syl_val
    workbook.save(file_path)


def run_classification(
    file_path: str,
    year: str,
    model_path: str,
    age_syl: List,
    matgen_syl: List,
    pupgen_syl: List,
    mother_syl: List,
    name_syl: List,
    sex_syl: List,
    session_syl: List,
    rec_num_syl: List,
    start_syl: List,
    end_syl: List,
    logger: Optional[logging.Logger] = None,
    year_audio_root: Optional[Path] = None,
    *,
    progress_hook: Optional[Callable[[int, int], None]] = None,
    save_npy: bool = True,
) -> Tuple[str, Optional[str]]:
    """Run the full syllable classification pipeline and write results to Excel.

    Orchestrates: load model, classify syllables, optional .npy dump, post-process,
    write 'Syllable number' column to the segmentation Excel file.

    Args:
        file_path: Path to the segmentation Excel file (will be updated)
        year: Recording year (used for building audio file paths)
        model_path: Path to the Keras model weights file (.h6)
        age_syl: Age values per syllable
        matgen_syl: Mother genotype values per syllable
        pupgen_syl: Pup genotype values per syllable
        mother_syl: Mother ID values per syllable
        name_syl: Pup name values per syllable
        sex_syl: Sex values per syllable
        session_syl: Session values per syllable
        rec_num_syl: Recording number values per syllable
        start_syl: Start time values per syllable
        end_syl: End time values per syllable
        logger: Optional logger instance

    Returns:
        Tuple of (output_xlsx_path, output_npy_path or None if save_npy is False)
    """
    if logger:
        logger.info("Classification started")

    model = load_classification_model(model_path)

    samples = classify_syllables(
        year,
        model,
        age_syl,
        matgen_syl,
        pupgen_syl,
        mother_syl,
        name_syl,
        sex_syl,
        session_syl,
        rec_num_syl,
        start_syl,
        end_syl,
        logger=logger,
        year_audio_root=year_audio_root,
        progress_hook=progress_hook,
    )
    if logger:
        logger.debug(f"Samples: {samples}")

    output_npy: Optional[str] = save_raw_predictions(file_path, samples) if save_npy else None

    syl_num = postprocess_predictions(samples, logger=logger)

    write_syllable_numbers(file_path, syl_num)

    if logger:
        logger.info(f"Classification finished (syllables={len(syl_num)})")

    return file_path, output_npy
