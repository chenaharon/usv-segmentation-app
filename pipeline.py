from __future__ import annotations

import gc
import os
import re
import sys
import time
from collections import deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

import librosa

from metadata_export import (
    MetadataRow,
    merge_syllable_counts_from_excel,
    save_metadata_inventory,
    save_processing_summary_workbook,
)


ProgressFn = Callable[..., None]


@dataclass
class RunSummary:
    """Aggregated run statistics for UI summary."""

    metadata_rows_scanned: int = 0
    wav_files_found: int = 0
    wav_segmentation_succeeded: int = 0
    wav_segmentation_failed: int = 0
    recordings_with_zero_syllables: int = 0
    total_syllable_rows: int = 0
    years_processed: List[str] = field(default_factory=list)
    output_files: List[str] = field(default_factory=list)
    error_messages: List[str] = field(default_factory=list)
    classification_model_path: Optional[str] = None
    output_directory: Optional[str] = None

    @property
    def recordings_with_syllables_in_output(self) -> int:
        """Successful segmentations that produced at least one syllable row in the workbook."""
        return max(0, self.wav_segmentation_succeeded - self.recordings_with_zero_syllables)

    @property
    def recordings_without_syllable_rows(self) -> int:
        """No rows in output: zero syllables after OK segmentation, or segmentation error."""
        return self.recordings_with_zero_syllables + self.wav_segmentation_failed

    def merge(self, other: "RunSummary") -> None:
        self.metadata_rows_scanned += other.metadata_rows_scanned
        self.wav_files_found += other.wav_files_found
        self.wav_segmentation_succeeded += other.wav_segmentation_succeeded
        self.wav_segmentation_failed += other.wav_segmentation_failed
        self.recordings_with_zero_syllables += other.recordings_with_zero_syllables
        self.total_syllable_rows += other.total_syllable_rows
        self.years_processed.extend(other.years_processed)
        self.output_files.extend(other.output_files)
        self.error_messages.extend(other.error_messages)
        if other.classification_model_path is not None:
            self.classification_model_path = other.classification_model_path
        if other.output_directory and not self.output_directory:
            self.output_directory = other.output_directory

    def format_report(self) -> str:
        lines = [
            "Run summary",
            f"  Years processed: {', '.join(self.years_processed) or '(none)'}",
            f"  Metadata rows scanned: {self.metadata_rows_scanned}",
            f"  Recording files resolved: {self.wav_files_found}",
            f"  Segmentation OK: {self.wav_segmentation_succeeded}",
            f"  Segmentation failed: {self.wav_segmentation_failed}",
            f"  Recordings with 0 syllables: {self.recordings_with_zero_syllables}",
            f"  Record files with syllables in output: {self.recordings_with_syllables_in_output}",
            f"  Record files with no syllable rows: {self.recordings_without_syllable_rows}",
            f"  Total syllable rows: {self.total_syllable_rows}",
        ]
        if self.classification_model_path:
            lines.append(f"  Classification model file: {self.classification_model_path}")
        if self.output_files:
            lines.append("  Outputs:")
            for p in self.output_files:
                lines.append(f"    - {p}")
        if self.error_messages:
            lines.append("  Notes:")
            for e in self.error_messages[:10]:
                lines.append(f"    - {e}")
            if len(self.error_messages) > 10:
                lines.append(f"    ... and {len(self.error_messages) - 10} more")
        return "\n".join(lines)


def _syllable_count_from_meta_row(r: MetadataRow) -> int:
    c = r.syllable_count
    if c is None:
        return 0
    try:
        return max(0, int(c))
    except (TypeError, ValueError):
        return 0


def _inventory_year_summary_row(inv: List[MetadataRow], year: str) -> Dict[str, Any]:
    """One summary row for ``save_processing_summary_workbook`` (English keys)."""
    y = str(year).strip()
    rows = [r for r in inv if str(r.year).strip() == y]
    if not rows:
        return {
            "Year": y,
            "Total mice (pups)": 0,
            "Total recordings": 0,
            "Total syllables": 0,
            "Mice with syllables detected": 0,
            "Recordings with syllables": 0,
        }
    from utils.audio_paths import pup_identity_key  # type: ignore

    def pup_key(r: MetadataRow) -> Optional[Tuple[str, str]]:
        n = str(r.name).strip()
        if not n:
            return None
        return (str(r.mother).strip().upper(), pup_identity_key(n))

    pups: Set[Tuple[str, str]] = set()
    pups_syl: Set[Tuple[str, str]] = set()
    for r in rows:
        k = pup_key(r)
        if k is not None:
            pups.add(k)
            if _syllable_count_from_meta_row(r) > 0:
                pups_syl.add(k)

    rec_syl = sum(1 for r in rows if _syllable_count_from_meta_row(r) > 0)
    total_syllables = sum(_syllable_count_from_meta_row(r) for r in rows)

    return {
        "Year": y,
        "Total mice (pups)": len(pups),
        "Total recordings": len(rows),
        "Total syllables": total_syllables,
        "Mice with syllables detected": len(pups_syl),
        "Recordings with syllables": rec_syl,
    }


@dataclass
class PipelineOptions:
    root_folder: str
    output_dir: Optional[str] = None
    years: Optional[List[str]] = None
    want_syllables_xlsx: bool = True
    want_metadata_xlsx: bool = True
    metadata_only: bool = False
    run_classification: bool = True
    # year (string) -> list of relative POSIX paths under that year folder; empty/omit = no filter
    subfolder_filters: Optional[Dict[str, List[str]]] = None
    # year (string) -> explicit metadata workbook path selected by the user
    metadata_file_overrides: Optional[Dict[str, str]] = None


def _emit_progress(
    fn: ProgressFn,
    p: float,
    msg: str,
    eta_seconds: Optional[float] = None,
) -> None:
    try:
        fn(p, msg, eta_seconds)
    except TypeError:
        fn(p, msg)


def output_timestamp_suffix(when: Optional[datetime] = None) -> str:
    """Human-readable timestamp for output filenames, e.g. 2026-03-28_19-30-53."""
    dt = when or datetime.now()
    return dt.strftime("%Y-%m-%d_%H-%M-%S")


def _is_per_year_segmentation_xlsx(path: str) -> bool:
    n = Path(path).name
    return (
        n.startswith("segmentation_")
        and n.endswith(".xlsx")
        and "Multiple_Years" not in n
    )


def _is_segmentation_workbook_path(path: str) -> bool:
    """Main segmentation output (excludes ``*_summary.xlsx`` workbooks)."""
    n = Path(path).name.lower()
    return n.startswith("segmentation_") and n.endswith(".xlsx") and not n.endswith("_summary.xlsx")


def _segmentation_workbook_for_summary_name(output_files: List[str]) -> Optional[Path]:
    """Latest segmentation workbook path to derive ``<stem>_summary.xlsx`` (after multi-year merge)."""
    for p in reversed(output_files):
        if _is_segmentation_workbook_path(p):
            return Path(p)
    return None


def _metadata_workbook_for_summary_name(output_files: List[str]) -> Optional[Path]:
    """Fallback when no segmentation file: pair summary name with last metadata inventory."""
    for p in reversed(output_files):
        path = Path(p)
        n = path.name.lower()
        if n.startswith("recordings_metadata_") and n.endswith(".xlsx") and not n.endswith("_summary.xlsx"):
            return path
    return None


def _ordered_unique_per_year_segmentation_paths(paths: List[str]) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for p in paths:
        if not _is_per_year_segmentation_xlsx(p):
            continue
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


def _renumber_index_column_in_worksheet(ws) -> None:
    """Set the ``Index`` column to 1..N for every data row (row 2 onward)."""
    if ws.max_row is None or ws.max_row < 2:
        return
    headers = [c.value for c in ws[1]]
    col_idx = None
    for i, h in enumerate(headers):
        if h is not None and str(h).strip().lower() == "index":
            col_idx = i + 1
            break
    if col_idx is None:
        return
    serial = 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col_idx).value = serial
        serial += 1


def _merge_segmentation_workbooks(source_paths: List[str], dest: Path) -> str:
    """
    Use the first workbook as the base (header + first year's rows), append data rows
    from the remaining files (from row 2 onward), renumber ``Index`` to one contiguous
    1-based series, then save to dest.
    """
    from openpyxl import load_workbook

    if len(source_paths) < 2:
        raise ValueError("merge requires at least two source workbooks")
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb_out = load_workbook(source_paths[0])
    ws_out = wb_out.active
    for src in source_paths[1:]:
        wb = load_workbook(src, read_only=True, data_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                cells = list(row)
                if not cells or all(v is None or v == "" for v in cells):
                    continue
                ws_out.append(cells)
        finally:
            wb.close()
    _renumber_index_column_in_worksheet(ws_out)
    wb_out.save(str(dest))
    return str(dest.resolve())


def _is_generated_pipeline_xlsx(path: Path) -> bool:
    """True if the filename looks like an app output workbook (not lab source metadata)."""
    stem_l = path.stem.lower()
    name_l = path.name.lower()
    return (
        stem_l.startswith("recordings_metadata_")
        or stem_l.startswith("segmentation_")
        or stem_l.startswith("processing_summary_")
        or name_l.endswith("_summary.xlsx")
    )


def _is_source_excel_file(path: Path) -> bool:
    if not path.is_file() or path.name.startswith("~$"):
        return False
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return False
    return not _is_generated_pipeline_xlsx(path)


_SKIP_DISCOVERY_DIR_NAMES = frozenset({"outputs", ".git", "__pycache__", "node_modules", ".venv"})


def _candidate_source_excels_under(selected: Path) -> List[Path]:
    """All ``.xlsx`` / ``.xls`` under *selected* (recursive), excluding temp and pipeline outputs."""
    sel = selected.resolve()
    found: List[Path] = []
    try:
        for pattern in ("*.xlsx", "*.xls"):
            for p in sel.rglob(pattern):
                if not p.is_file() or p.name.startswith("~$"):
                    continue
                if _is_generated_pipeline_xlsx(p):
                    continue
                try:
                    rel = p.relative_to(sel)
                except ValueError:
                    continue
                if any(part in _SKIP_DISCOVERY_DIR_NAMES for part in rel.parts):
                    continue
                found.append(p)
    except OSError:
        pass
    return found


def _candidate_source_excels_near_year(selected: Path) -> List[Path]:
    """
    Candidate Excel files under the year folder itself or exactly one subdirectory below it.
    """
    sel = selected.resolve()
    found: List[Path] = []
    try:
        for pattern in ("*.xlsx", "*.xls"):
            for p in sel.glob(pattern):
                if _is_source_excel_file(p):
                    found.append(p)
            for d in sel.iterdir():
                if not d.is_dir() or d.name in _SKIP_DISCOVERY_DIR_NAMES:
                    continue
                for p in d.glob(pattern):
                    if _is_source_excel_file(p):
                        found.append(p)
    except OSError:
        pass
    return sorted(set(found), key=lambda p: str(p).lower())


def year_metadata_availability(year_path: Path) -> bool:
    """
    UI hint: True if the year folder contains at least one source Excel file (``.xlsx`` / ``.xls``).

    Checks recursively under the year directory (skips ``outputs/``, ``.git``, etc.).
    Ignores Excel temp files (``~$``) and known pipeline outputs (``segmentation_*``, ``recordings_metadata_*``).
    This does not validate column headers; the pipeline still enforces that at runtime.
    """
    year_path = year_path.resolve()
    for p in _candidate_source_excels_near_year(year_path):
        if _is_source_excel_file(p):
            return True
    return False


def _normalize_subfolder_prefix(p: str) -> str:
    s = (p or "").strip().replace("\\", "/")
    while s.startswith("./"):
        s = s[2:]
    s = s.strip("/")
    return s.lower()


def _rel_paths_for_subfolder_filter(
    wav: Path,
    year_root: Path,
    dataset_root: Optional[Path],
    year: str,
) -> List[str]:
    """Relative path strings to match user-selected subfolder prefixes."""
    w = wav.resolve()
    out: List[str] = []
    try:
        out.append(w.relative_to(year_root.resolve()).as_posix())
    except ValueError:
        pass
    if dataset_root is not None:
        dr = dataset_root.resolve()
        try:
            out.append(w.relative_to(dr).as_posix())
        except ValueError:
            pass
        anchor = (dr / "USV_Recordings" / year).resolve()
        try:
            out.append(w.relative_to(anchor).as_posix())
        except ValueError:
            pass
    return out


def wav_matches_subfolder_prefixes(
    wav: Path,
    year_root: Path,
    dataset_root: Optional[Path],
    year: str,
    prefixes: Optional[List[str]],
) -> bool:
    """
    * ``prefixes is None``: no restriction (process all).
    * ``prefixes == []``: explicit exclusion (match nothing).
    * Otherwise the recording file must lie under the year folder such that at least one relative
      path matches a normalized prefix (folder tree selection in the UI).
    """
    if prefixes is None:
        return True
    if len(prefixes) == 0:
        return False
    norm_pref = [
        _normalize_subfolder_prefix(p) for p in prefixes if _normalize_subfolder_prefix(p)
    ]
    if not norm_pref:
        return False
    rels = _rel_paths_for_subfolder_filter(wav, year_root, dataset_root, year)
    if not rels:
        return False
    for rel in rels:
        r = rel.lower()
        for p in norm_pref:
            if r == p or r.startswith(p + "/"):
                return True
    return False


def _wav_log_path(wav: Path) -> str:
    """Full path for progress logs (basenames repeat across folders)."""
    try:
        return str(wav.resolve())
    except OSError:
        return str(wav)


def _opt_year_subfolders(
    filters: Optional[Dict[str, List[str]]], year_str: str
) -> Optional[List[str]]:
    """Return prefix list for *year_str* (may be ``[]``). None = no subfolder restriction."""
    if not filters:
        return None
    if year_str not in filters:
        return None
    return filters[year_str]


def _runtime_base_dir() -> Path:
    frozen_base = getattr(sys, "_MEIPASS", None)
    if frozen_base:
        return Path(frozen_base)
    return Path(__file__).resolve().parent


def _add_preprocessing_to_path(preprocessing_dir: Path) -> None:
    path_str = str(preprocessing_dir)
    if path_str not in sys.path:
        sys.path.insert(0, path_str)


def _scan_audio_files(
    root: Path,
    *,
    progress: Optional[ProgressFn] = None,
    progress_p: float = 0.05,
    year: str = "",
    dataset_root: Optional[Path] = None,
    subfolder_prefixes: Optional[List[str]] = None,
) -> List[Path]:
    """Collect supported recording files (currently ``.wav``/``.wave``) under ``root``; optional progress while walking.

    When ``subfolder_prefixes`` is set (including ``[]``), paths are restricted with
    ``wav_matches_subfolder_prefixes`` before the progress line so the count matches
    the UI folder selection, not the whole year tree.
    """
    wav_ext = {".wav", ".wave"}
    found: List[Path] = []
    prefix = f"[{year}] " if year else ""
    for dirpath, _dirnames, filenames in os.walk(root):
        for fn in filenames:
            if Path(fn).suffix.lower() not in wav_ext:
                continue
            found.append(Path(dirpath) / fn)

    root_resolved = root.resolve()
    scanned_n = len(found)
    if subfolder_prefixes is not None:
        found = [
            p
            for p in found
            if wav_matches_subfolder_prefixes(
                p.resolve(), root_resolved, dataset_root, year, subfolder_prefixes
            )
        ]

    if progress is not None:
        if subfolder_prefixes is None:
            msg = f"{prefix}{len(found)} audio recordings found"
        elif scanned_n != len(found):
            msg = (
                f"{prefix}{len(found)} audio recordings found under selected folder(s) "
                f"({scanned_n} total under year folder)"
            )
        else:
            msg = f"{prefix}{len(found)} audio recordings found under selected folder(s)"
        _emit_progress(progress, progress_p, msg, None)
    return sorted(found, key=lambda p: str(p).lower())


def _year_folder_name(selected: Path) -> str:
    if selected.name.isdigit() and len(selected.name) == 4:
        return selected.name
    m = re.search(r"(19|20)\d{2}", str(selected))
    return m.group(0) if m else selected.name


def discover_year_roots(user_root: Path) -> List[Tuple[str, Path]]:
    """
    Return list of (year_string, year_folder_path).

    If the selected folder is itself a 4-digit year, return a single pair.
    Otherwise scan immediate subdirectories named as four digits.
    If none match, treat the whole folder as one synthetic year from its name.
    """
    user_root = user_root.resolve()
    if user_root.name.isdigit() and len(user_root.name) == 4:
        return [(user_root.name, user_root)]
    found: List[Tuple[str, Path]] = []
    try:
        for p in sorted(user_root.iterdir()):
            if p.is_dir() and re.fullmatch(r"\d{4}", p.name):
                found.append((p.name, p.resolve()))
    except OSError:
        pass
    if found:
        return found
    return [(_year_folder_name(user_root), user_root)]


def _is_year_only_folder(path: Path) -> bool:
    return path.name.isdigit() and len(path.name) == 4


def _find_metadata_workbooks_client_layout(
    selected: Path,
    explicit_metadata_file: Optional[Path] = None,
) -> List[Path]:
    """Discover metadata workbooks under year root / one level below, or use explicit path."""
    sel = selected.resolve()
    out: List[Path] = []
    if explicit_metadata_file is not None:
        p = explicit_metadata_file.resolve()
        if _is_valid_metadata_workbook(p):
            return [p]
        return []
    for p in _candidate_source_excels_near_year(sel):
        if _is_valid_metadata_workbook(p):
            out.append(p)
    return sorted(set(out), key=lambda p: str(p).lower())


def _is_valid_metadata_workbook(path: Path) -> bool:
    try:
        import pandas as pd
        from utils.io_utils import metadata_columns_satisfied  # type: ignore

        try:
            df = pd.read_excel(path, sheet_name=0, nrows=0, engine="openpyxl")
        except Exception:
            df = pd.read_excel(path, sheet_name=0, nrows=0, engine="xlrd")
        cols = {str(c).strip() for c in df.columns}
        return metadata_columns_satisfied(cols)
    except Exception:
        return False


def _is_pup_summary_workbook(path: Path) -> bool:
    """True for pup tables with Mother + Name + Gender (USV pups / טבלת עכברים style)."""
    try:
        import pandas as pd
        from utils.io_utils import pup_summary_columns_satisfied  # type: ignore

        try:
            df = pd.read_excel(path, sheet_name=0, nrows=0, engine="openpyxl")
        except Exception:
            df = pd.read_excel(path, sheet_name=0, nrows=0, engine="xlrd")
        cols = {str(c).strip() for c in df.columns}
        return pup_summary_columns_satisfied(cols)
    except Exception:
        return False


def _find_pup_summary_workbooks(
    selected: Path,
    explicit_metadata_file: Optional[Path] = None,
) -> List[Path]:
    """Pup-summary xlsx under year root / one level below, or explicit path if provided."""
    sel = selected.resolve()
    out: List[Path] = []
    if explicit_metadata_file is not None:
        p = explicit_metadata_file.resolve()
        if _is_pup_summary_workbook(p):
            return [p]
        return []
    for p in _candidate_source_excels_near_year(sel):
        if _is_pup_summary_workbook(p):
            out.append(p)
    return sorted(set(out), key=lambda p: str(p).lower())


def _merge_sex_lookups_from_year_folder(
    selected: Path,
    explicit_metadata_file: Optional[Path] = None,
) -> Dict[Tuple[str, str], str]:
    from utils.io_utils import build_sex_lookup_from_pup_summary_xlsx  # type: ignore

    merged: Dict[Tuple[str, str], str] = {}
    for p in _find_pup_summary_workbooks(selected, explicit_metadata_file):
        try:
            merged.update(build_sex_lookup_from_pup_summary_xlsx(str(p)))
        except Exception:
            continue
    return merged


def _merge_pup_details_from_year_folder(
    selected: Path,
    explicit_metadata_file: Optional[Path] = None,
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    from utils.io_utils import build_pup_summary_details_lookup_xlsx  # type: ignore

    merged: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for p in _find_pup_summary_workbooks(selected, explicit_metadata_file):
        try:
            merged.update(build_pup_summary_details_lookup_xlsx(str(p)))
        except Exception:
            continue
    return merged


def _resolve_sex_from_pup_tables(
    mother: str,
    name_from_path: str,
    lookup: Dict[Tuple[str, str], str],
    *,
    extra_name_hints: Optional[List[str]] = None,
) -> str:
    from utils.audio_paths import iter_pup_table_lookup_keys  # type: ignore

    if not lookup:
        return _guess_sex(name_from_path)
    hints = [name_from_path] + list(extra_name_hints or [])
    for k in iter_pup_table_lookup_keys(mother, *hints):
        if k in lookup:
            return lookup[k]
    return _guess_sex(name_from_path)


def _excel_sex_with_pup_fallback(
    excel_sx: str,
    mother: str,
    name_key: str,
    lookup: Dict[Tuple[str, str], str],
    *,
    extra_name_hints: Optional[List[str]] = None,
) -> str:
    from utils.io_utils import normalize_sex_cell  # type: ignore

    sx = normalize_sex_cell(excel_sx)
    if sx != "U":
        return sx
    return _resolve_sex_from_pup_tables(
        mother, name_key, lookup, extra_name_hints=extra_name_hints
    )


def _pup_path_hints_for_wav(selected: Path, wav: Path) -> List[str]:
    raw = _raw_pup_folder_from_path(selected, wav)
    if not raw:
        return []
    from utils.audio_paths import canonical_pup_display_name  # type: ignore

    cd = canonical_pup_display_name(raw)
    out = [raw]
    if cd and cd != raw:
        out.append(cd)
    return out


def _output_pup_name_for_segmentation(selected: Path, wav: Path, excel_name: str) -> str:
    """Prefer folder-based canonical name; fall back to metadata Name."""
    from utils.audio_paths import canonical_pup_display_name  # type: ignore

    raw = _raw_pup_folder_from_path(selected, wav)
    if raw:
        c = canonical_pup_display_name(raw)
        if c:
            return c
    c2 = canonical_pup_display_name(excel_name)
    return c2 or str(excel_name).strip() or "UnknownName"


def _sex_for_metadata_row(
    excel_sx: str,
    mother: str,
    excel_name: str,
    lookup: Dict[Tuple[str, str], str],
    selected: Path,
    wav: Optional[Path],
) -> str:
    """Resolve sex using Excel cell first, then pup tables with path-aware hints."""
    from utils.io_utils import normalize_sex_cell  # type: ignore

    sx0 = normalize_sex_cell(excel_sx)
    if sx0 != "U":
        return sx0
    hints = _pup_path_hints_for_wav(selected, wav) if wav is not None else []
    return _resolve_sex_from_pup_tables(
        mother, excel_name, lookup, extra_name_hints=hints
    )


def _resolve_dataset_and_metadata(selected: Path) -> Tuple[Optional[Path], Optional[Path]]:
    sel = selected.resolve()
    if sel.parent.name.lower() == "usv_recordings":
        root = sel.parent.parent
        meta = root / "metadata"
        if meta.is_dir():
            return root, meta
    if (sel / "metadata").is_dir() and (sel / "USV_Recordings").is_dir():
        return sel, sel / "metadata"
    if (sel.parent / "metadata").is_dir() and (sel.parent / "USV_Recordings").is_dir():
        return sel.parent, sel.parent / "metadata"
    return None, None


def _load_merged_metadata(metadata_dir: Path, year_hint: str) -> Dict[str, List]:
    from utils import (  # type: ignore
        METADATA_REQUIRED_COLUMNS,
        extract_year_from_filename,
        read_metadata_as_lists,
    )

    files = sorted(
        p
        for p in metadata_dir.iterdir()
        if p.is_file() and p.suffix.lower() in (".xlsx", ".xls") and not p.name.startswith("~$")
    )
    if not files:
        return {}

    matching: List[Path] = []
    for p in files:
        try:
            if extract_year_from_filename(p.name) == year_hint:
                matching.append(p)
        except ValueError:
            continue
    use_files = matching if matching else files

    merged: Dict[str, List] = {c: [] for c in METADATA_REQUIRED_COLUMNS}
    for p in use_files:
        data = read_metadata_as_lists(str(p))
        for c in METADATA_REQUIRED_COLUMNS:
            merged[c].extend(data[c])
    return merged


def _meta_row(meta: Dict[str, List], i: int) -> Tuple[str, str, str, str, str, int, int, str]:
    def g(col: str) -> Any:
        return meta[col][i]

    mother = str(g("Mother")).strip()
    matgen = str(g("Mother Genotype")).strip()
    name = str(g("Name")).strip()
    sex = str(g("Sex")).strip()
    pupgen = str(g("Offspring Genotype")).strip()
    day = _to_int(g("Day"))
    session = _normalize_session_value(_to_int(g("Session")))
    rec = g("Recording Number")
    rec_str = str(rec).strip()
    if rec_str.endswith(".0") and rec_str.replace(".0", "").isdigit():
        rec_str = rec_str[:-2]
    return mother, matgen, name, sex, pupgen, day, session, rec_str


def _to_int(v: Any) -> int:
    if v is None or (isinstance(v, float) and str(v) == "nan"):
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return 0


def _normalize_session_value(session: int) -> int:
    """Use 1 when Session is missing, unparsed, or non-positive (Day still defaults via _to_int → 0)."""
    return session if session > 0 else 1


def _excel_path_from_resolved_wav(dataset_root: Path, wav: Path, year: str) -> str:
    anchor = (dataset_root / "USV_Recordings" / year).resolve()
    try:
        rel = wav.resolve().relative_to(anchor)
        return f"USV_Recordings/{year}/{rel.as_posix()}"
    except ValueError:
        return f"USV_Recordings/{year}/{wav.name}"


def _excel_path_column(selected: Path, wav_path: Path, year: str) -> str:
    rel = wav_path.relative_to(selected).as_posix()
    return f"USV_Recordings/{year}/{rel}"


def _resolve_model_path() -> Tuple[Optional[Path], List[str]]:
    """
    Find ``model_weights.h6`` regardless of ``os.getcwd()`` (e.g. user output under Downloads).

    Override with env ``USV_MODEL_PATH`` pointing to the file or its parent directory.

    Returns:
        (resolved file path or None, unique list of candidate paths checked, for diagnostics).
    """
    base_dir = _runtime_base_dir()
    app_root = Path(__file__).resolve().parent

    candidates: List[Path] = []

    env = os.environ.get("USV_MODEL_PATH", "").strip()
    if env:
        p = Path(env).expanduser()
        candidates.append(p)
        if p.is_dir():
            candidates.append(p / "model_weights.h6")

    candidates.extend(
        [
            Path(sys.executable).resolve().parent / "models" / "model_weights.h6",
            app_root / "models" / "model_weights.h6",
            Path.cwd() / "models" / "model_weights.h6",
            base_dir / "models" / "model_weights.h6",
            base_dir / "preprocessing" / "src" / "models" / "model_weights.h6",
        ]
    )

    tried: List[str] = []
    seen: set[str] = set()
    for path in candidates:
        try:
            label = str(path.expanduser().resolve())
        except OSError:
            label = str(path.expanduser())
        if label not in seen:
            seen.add(label)
            tried.append(label)

    def _is_usable_keras_model_path(p: Path) -> bool:
        q = p.expanduser()
        if not q.exists():
            return False
        if q.is_file():
            return True
        # SavedModel bundle (e.g. folder ``model_weights.h6/`` with saved_model.pb + variables/)
        if q.is_dir() and (q / "saved_model.pb").is_file():
            return True
        return False

    for path in candidates:
        if _is_usable_keras_model_path(path):
            return path.expanduser().resolve(), tried
    return None, tried


def _write_constant_syllable_column(file_path: str, value: int) -> None:
    import pandas as pd

    df = pd.read_excel(file_path, sheet_name=0, engine="openpyxl")
    df["Syllable number"] = int(value)
    df.to_excel(file_path, index=False, engine="openpyxl")


def _extract_row_metadata_from_path_layout(
    root: Path, wav_path: Path
) -> Tuple[str, str, str, str, str, int, int, str]:
    folders = _extract_layout_folders(root, wav_path)
    stem = wav_path.stem

    mother, matgen = "UnknownMother", "UNK"
    name, pupgen = "UnknownName", "UNK"
    day, session = 0, 0

    mother_idx = -1
    for i, token in enumerate(folders):
        if _is_mother_folder_token(token):
            mother_idx = i
            break
    if mother_idx >= 0:
        mother, matgen = _parse_mother_from_folder_token(folders[mother_idx])
        if mother_idx + 1 < len(folders):
            name, pupgen = _parse_pup_from_folder_token(folders[mother_idx + 1])
        for token in folders[mother_idx + 1 :]:
            low = token.lower().replace(" ", "_")
            if day == 0 and low.startswith("day"):
                day = _extract_number(token)
            if session == 0 and low.startswith("session"):
                session = _extract_number(token)
        rec_num = stem
        sex = _guess_sex(name)
        return mother, matgen, name, sex, pupgen, day, _normalize_session_value(session), rec_num

    if len(folders) >= 4:
        mother, matgen = _split_pair(folders[0], "UnknownMother", "UNK")
        name, pupgen = _split_pair(folders[1], "UnknownName", "UNK")
        name = _normalize_pup_name_from_folder(name)
        day = _extract_number(folders[2])
        session = _extract_number(folders[3])
    elif len(folders) == 3:
        mother, matgen = _split_pair(folders[0], "UnknownMother", "UNK")
        if folders[1].lower().startswith("day"):
            day = _extract_number(folders[1])
            session = _extract_number(folders[2])
        else:
            name, pupgen = _split_pair(folders[1], "UnknownName", "UNK")
            name = _normalize_pup_name_from_folder(name)
            day = _extract_number(folders[2])
    elif len(folders) == 2:
        mother, matgen = _split_pair(folders[0], "UnknownMother", "UNK")
        if folders[1].lower().startswith("day"):
            day = _extract_number(folders[1])
        elif folders[1].lower().startswith("session"):
            session = _extract_number(folders[1])
        else:
            name, pupgen = _split_pair(folders[1], "UnknownName", "UNK")
            name = _normalize_pup_name_from_folder(name)
    elif len(folders) == 1:
        mother, matgen = _split_pair(folders[0], "UnknownMother", "UNK")

    rec_num = stem
    sex = _guess_sex(name)
    return mother, matgen, name, sex, pupgen, day, _normalize_session_value(session), rec_num


def _split_pair(value: str, default_left: str, default_right: str) -> Tuple[str, str]:
    raw = str(value).strip()
    if not raw:
        return default_left, default_right
    parts = [p for p in re.split(r"[_\s]+", raw) if p]
    if len(parts) >= 2:
        geno = _extract_genotype_from_parts(parts)
        return parts[0] or default_left, geno or default_right
    if "_" in raw:
        left, right = raw.split("_", 1)
        return left or default_left, _normalize_genotype_token(right) or default_right
    return raw or default_left, default_right


def _normalize_genotype_token(token: str) -> str:
    t = str(token).strip().upper().replace("-", "").replace("_", "")
    if t in {"HET", "HT"}:
        return "HT"
    if t in {"WT", "WILDTYPE"} or re.fullmatch(r"(WT)+", t):
        return "WT"
    if re.fullmatch(r"(HET)+", t) or re.fullmatch(r"(HT)+", t):
        return "HT"
    if t in {"HOM", "KO", "KI"}:
        return t
    return t or "UNK"


def _extract_genotype_from_parts(parts: List[str]) -> Optional[str]:
    """Find first valid genotype token (WT/HT/HOM/KO/KI) in tokenized folder labels."""
    for p in parts:
        g = _normalize_genotype_token(p)
        if g in {"WT", "HT", "HOM", "KO", "KI"}:
            return g
    return None


def _normalize_pup_name_from_folder(value: str) -> str:
    from utils.audio_paths import canonical_pup_display_name  # type: ignore

    c = canonical_pup_display_name(value)
    return c if c else "UnknownName"


def _is_mother_folder_token(value: str) -> bool:
    s = str(value).strip()
    if not s:
        return False
    parts = [p for p in re.split(r"[_\s]+", s) if p]
    if len(parts) < 2:
        return False
    return _extract_genotype_from_parts(parts) in {"WT", "HT", "HOM", "KO", "KI"}


def _parse_mother_from_folder_token(value: str) -> Tuple[str, str]:
    parts = [p for p in re.split(r"[_\s]+", str(value).strip()) if p]
    if len(parts) >= 2:
        return parts[0], (_extract_genotype_from_parts(parts) or "UNK")
    return _split_pair(str(value), "UnknownMother", "UNK")


def _parse_pup_from_folder_token(value: str) -> Tuple[str, str]:
    from utils.audio_paths import canonical_pup_display_name  # type: ignore

    s = str(value).strip()
    if not s:
        return "UnknownName", "UNK"
    parts = [p for p in re.split(r"[_\s]+", s) if p]
    geno = _extract_genotype_from_parts(parts)
    if geno is not None:
        name_parts = [
            p for p in parts if _normalize_genotype_token(p) not in {"WT", "HT", "HOM", "KO", "KI"}
        ]
        base = "_".join(name_parts) if name_parts else s
        return canonical_pup_display_name(base) or "UnknownName", geno
    return canonical_pup_display_name(s) or "UnknownName", "UNK"


def _extract_layout_folders(root: Path, wav_path: Path) -> List[str]:
    rel = wav_path.relative_to(root)
    parts = list(rel.parts)
    folders = parts[:-1] if len(parts) > 1 else []
    if len(folders) >= 5 and re.fullmatch(r"\d{4}", str(folders[0])):
        folders = folders[1:]
    return [str(x) for x in folders]


def _raw_pup_folder_from_path(root: Path, wav_path: Path) -> str:
    folders = _extract_layout_folders(root, wav_path)
    if not folders:
        return ""
    mother_idx = -1
    for i, token in enumerate(folders):
        if _is_mother_folder_token(token):
            mother_idx = i
            break
    if mother_idx >= 0 and mother_idx + 1 < len(folders):
        return folders[mother_idx + 1]
    if len(folders) >= 2:
        return folders[1]
    return ""


def _resolve_pup_details(
    mother: str,
    raw_pup_folder: str,
    normalized_name: str,
    details_lookup: Dict[Tuple[str, str], Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    from utils.audio_paths import iter_pup_table_lookup_keys  # type: ignore

    if not details_lookup:
        return None
    hints = [str(raw_pup_folder).strip(), str(normalized_name).strip()]
    for k in iter_pup_table_lookup_keys(mother, *hints):
        if k in details_lookup:
            return details_lookup[k]
    return None


def _extract_number(value: str) -> int:
    m = re.search(r"\d+", str(value))
    return int(m.group(0)) if m else 0


def _guess_sex(name: str) -> str:
    n = name.upper()
    if n.endswith("M") or "_M" in n or "MALE" in n:
        return "M"
    if n.endswith("F") or "_F" in n or "FEMALE" in n:
        return "F"
    return "U"


def _classification_year_root(
    selected: Path,
    year: str,
    dataset_root: Optional[Path],
) -> Path:
    """Directory that contains ``Mother_* / Name_* / day_* / session* / recording file``."""
    if dataset_root is not None:
        p = (dataset_root / "USV_Recordings" / year).resolve()
        if p.is_dir():
            return p
    return selected.resolve()


def _resolve_wav_under_mother_folder(
    mother_dir: Path,
    m: str,
    mg: str,
    n_: str,
    pg: str,
    d: int,
    ses: int,
    rec: str,
    resolve_wav_path_fn,
) -> Optional[Path]:
    from utils.audio_paths import iter_recording_base_candidates  # type: ignore

    for base in iter_recording_base_candidates(
        [mother_dir],
        m,
        mg,
        n_,
        pg,
        d,
        ses,
        rec,
        nested_mother_folder=True,
    ):
        w = resolve_wav_path_fn(base)
        if w is not None:
            return w.resolve()
    return None


def _count_segmentable_client_metadata_rows(
    selected: Path,
    year: str,
    client_metadata_files: List[Path],
    dataset_root: Optional[Path],
    subfolder_prefixes: Optional[List[str]],
    sex_lookup: Dict[Tuple[str, str], str],
    resolve_wav_path_fn,
) -> int:
    """Rows that resolve to a recording file under the year and pass the subfolder filter (for ETA)."""
    from utils import read_metadata_as_lists  # type: ignore
    from utils.audio_paths import resolve_wav_under_year_folder  # type: ignore

    sel_resolved = selected.resolve()
    n = 0
    for mf in client_metadata_files:
        try:
            data = read_metadata_as_lists(str(mf))
        except Exception:
            continue
        for i in range(len(data["Mother"])):
            m, mg, n_, _, pg, d, ses, rec = _meta_row(data, i)
            if mf.parent.resolve() == sel_resolved:
                wav = resolve_wav_under_year_folder(
                    sel_resolved, m, mg, n_, pg, d, ses, rec
                )
            else:
                wav = _resolve_wav_under_mother_folder(
                    mf.parent, m, mg, n_, pg, d, ses, rec, resolve_wav_path_fn
                )
            if wav is None:
                continue
            wav = wav.resolve()
            if _is_year_only_folder(selected):
                try:
                    wav.relative_to(sel_resolved)
                except ValueError:
                    continue
            if not wav_matches_subfolder_prefixes(
                wav, selected, dataset_root, year, subfolder_prefixes
            ):
                continue
            n += 1
    return n


def _count_segmentable_usv_metadata_rows(
    selected: Path,
    year: str,
    meta: Dict[str, List],
    dataset_root: Optional[Path],
    subfolder_prefixes: Optional[List[str]],
    sex_lookup: Dict[Tuple[str, str], str],
) -> int:
    from utils.audio_paths import resolve_wav_usv_recordings_layout  # type: ignore

    n = 0
    for i in range(len(meta.get("Mother", []))):
        m, mg, n_, _, pg, d, ses, rec = _meta_row(meta, i)
        wav = resolve_wav_usv_recordings_layout(
            "USV_Recordings", year, m, mg, n_, pg, int(d), int(ses), rec
        )
        if wav is None:
            continue
        wav = wav.resolve()
        if _is_year_only_folder(selected):
            try:
                wav.relative_to(selected.resolve())
            except ValueError:
                continue
        if not wav_matches_subfolder_prefixes(
            wav, selected, dataset_root, year, subfolder_prefixes
        ):
            continue
        n += 1
    return n


def collect_metadata_inventory_only(
    *,
    selected: Path,
    year: str,
    resolve_wav_path_fn,
    progress: ProgressFn,
    progress_base: float,
    progress_span: float,
    subfolder_prefixes: Optional[List[str]] = None,
    explicit_metadata_file: Optional[Path] = None,
) -> Tuple[List[MetadataRow], RunSummary]:
    from utils import read_metadata_as_lists  # type: ignore
    from utils.audio_paths import resolve_wav_under_year_folder  # type: ignore

    summary = RunSummary()
    rows: List[MetadataRow] = []
    sel_resolved = selected.resolve()
    client_metadata_files = _find_metadata_workbooks_client_layout(selected, explicit_metadata_file)
    dataset_root, metadata_dir = _resolve_dataset_and_metadata(selected)
    sex_lookup = _merge_sex_lookups_from_year_folder(selected, explicit_metadata_file)
    pup_details_lookup = _merge_pup_details_from_year_folder(selected, explicit_metadata_file)

    def add_row(
        mf_name: str,
        m: str,
        mg: str,
        n_: str,
        sx: str,
        pg: str,
        d: int,
        ses: int,
        rec: str,
        wav: Optional[Path],
        status: str,
        *,
        count_as_found: bool = True,
    ) -> None:
        summary.metadata_rows_scanned += 1
        abs_p = str(wav.resolve()) if wav is not None else ""
        path_style = _excel_path_column(selected, wav, year) if wav is not None else ""
        if wav is not None and count_as_found:
            summary.wav_files_found += 1
        rows.append(
            MetadataRow(
                year=year,
                metadata_file=mf_name,
                mother=m,
                mother_genotype=mg,
                name=n_,
                sex=sx,
                offspring_genotype=pg,
                day=d,
                session=ses,
                recording_number=rec,
                wav_absolute_path=abs_p,
                path_column_style=path_style,
                status=status,
            )
        )

    total_rows = 0
    if client_metadata_files:
        for mf in client_metadata_files:
            try:
                data = read_metadata_as_lists(str(mf))
                total_rows += len(data["Mother"])
            except Exception:
                continue
    elif dataset_root is not None and metadata_dir is not None:
        meta = _load_merged_metadata(metadata_dir, year)
        total_rows = len(meta.get("Mother", []))

    done = 0
    if client_metadata_files:
        for mf in client_metadata_files:
            try:
                data = read_metadata_as_lists(str(mf))
            except Exception:
                continue
            n = len(data["Mother"])
            for i in range(n):
                m, mg, n_, sx, pg, d, ses, rec = _meta_row(data, i)
                if mf.parent.resolve() == sel_resolved:
                    wav = resolve_wav_under_year_folder(
                        sel_resolved, m, mg, n_, pg, d, ses, rec
                    )
                else:
                    wav = _resolve_wav_under_mother_folder(
                        mf.parent, m, mg, n_, pg, d, ses, rec, resolve_wav_path_fn
                    )
                if wav is None:
                    sx_o = _excel_sex_with_pup_fallback(sx, m, n_, sex_lookup)
                    add_row(mf.name, m, mg, n_, sx_o, pg, d, ses, rec, None, "Recording file not found")
                else:
                    wav = wav.resolve()
                    sx_o = _sex_for_metadata_row(sx, m, n_, sex_lookup, sel_resolved, wav)
                    disp = _output_pup_name_for_segmentation(sel_resolved, wav, n_)
                    if _is_year_only_folder(selected):
                        try:
                            wav.relative_to(sel_resolved)
                        except ValueError:
                            done += 1
                            continue
                    if not wav_matches_subfolder_prefixes(
                        wav, selected, dataset_root, year, subfolder_prefixes
                    ):
                        add_row(
                            mf.name,
                            m,
                            mg,
                            disp,
                            sx_o,
                            pg,
                            d,
                            ses,
                            rec,
                            wav,
                            "Skipped (subfolder filter)",
                            count_as_found=False,
                        )
                    else:
                        add_row(
                            mf.name, m, mg, disp, sx_o, pg, d, ses, rec, wav, "Found"
                        )
                done += 1
                if total_rows > 0:
                    p = progress_base + progress_span * (done / total_rows)
                    _emit_progress(progress, p, f"[{year}] metadata scan {done}/{total_rows}")
    elif dataset_root is not None and metadata_dir is not None:
        from utils.audio_paths import resolve_wav_usv_recordings_layout  # type: ignore

        old_cwd = os.getcwd()
        os.chdir(str(dataset_root))
        try:
            meta = _load_merged_metadata(metadata_dir, year)
            n = len(meta.get("Mother", []))
            yr_root = selected.resolve()
            for i in range(n):
                m, mg, n_, sx, pg, d, ses, rec = _meta_row(meta, i)
                wav = resolve_wav_usv_recordings_layout(
                    "USV_Recordings", year, m, mg, n_, pg, int(d), int(ses), rec
                )
                if wav is None:
                    sx_o = _excel_sex_with_pup_fallback(sx, m, n_, sex_lookup)
                    add_row(
                        "(metadata dir)", m, mg, n_, sx_o, pg, d, ses, rec, None, "Recording file not found"
                    )
                else:
                    wav = wav.resolve()
                    sx_o = _sex_for_metadata_row(sx, m, n_, sex_lookup, yr_root, wav)
                    disp = _output_pup_name_for_segmentation(yr_root, wav, n_)
                    if _is_year_only_folder(selected):
                        try:
                            wav.relative_to(sel_resolved)
                        except ValueError:
                            done += 1
                            continue
                    if not wav_matches_subfolder_prefixes(
                        wav, selected, dataset_root, year, subfolder_prefixes
                    ):
                        add_row(
                            "(metadata dir)",
                            m,
                            mg,
                            disp,
                            sx_o,
                            pg,
                            d,
                            ses,
                            rec,
                            wav,
                            "Skipped (subfolder filter)",
                            count_as_found=False,
                        )
                    else:
                        add_row(
                            "(metadata dir)",
                            m,
                            mg,
                            disp,
                            sx_o,
                            pg,
                            d,
                            ses,
                            rec,
                            wav,
                            "Found",
                        )
                done += 1
                if n > 0:
                    p = progress_base + progress_span * (done / n)
                    _emit_progress(progress, p, f"[{year}] metadata scan {done}/{n}")
        finally:
            os.chdir(old_cwd)

    return rows, summary


def process_single_year(
    *,
    selected: Path,
    year: str,
    outputs_dir: Path,
    progress: ProgressFn,
    progress_lo: float,
    progress_hi: float,
    subfolder_prefixes: Optional[List[str]] = None,
    explicit_metadata_file: Optional[Path] = None,
    run_classification: bool = True,
) -> Tuple[Optional[str], RunSummary, List[MetadataRow]]:
    """
    Run segmentation + features + classification + enrich for one year folder.
    Returns (syllable xlsx path or None, summary, inventory rows with syllable counts filled when applicable).

    If ``subfolder_prefixes`` is non-empty, only recordings whose audio file path lies under
    one of those relative paths (under the year folder) are segmented. Metadata
    workbooks are still read in full; filtered rows are marked skipped in inventory.
    """
    from utils.audio_paths import resolve_wav_under_year_folder  # type: ignore
    from steps.segmentation import (  # type: ignore
        create_segmentation_workbook,
        segment_single_recording,
        FRAME_LENGTH,
        OVERLAP,
        THRESH,
        HARMONY_TH,
    )
    from steps.read_segmentation import read_segmentation_results  # type: ignore
    from steps.compute_basic_features import compute_basic_features  # type: ignore
    from steps.classification import run_classification as run_syllable_classification  # type: ignore
    from steps.enrich_columns import enrich_segmentation_columns  # type: ignore
    from utils import read_metadata_as_lists  # type: ignore

    summary = RunSummary()
    inventory: List[MetadataRow] = []

    def span_t(p: float) -> float:
        return progress_lo + (progress_hi - progress_lo) * max(0.0, min(1.0, p))

    dataset_root, metadata_dir = _resolve_dataset_and_metadata(selected)
    client_metadata_files = _find_metadata_workbooks_client_layout(selected, explicit_metadata_file)
    sex_lookup = _merge_sex_lookups_from_year_folder(selected, explicit_metadata_file)
    pup_details_lookup = _merge_pup_details_from_year_folder(selected, explicit_metadata_file)

    book, sheet = create_segmentation_workbook()
    audio_paths: List[Path] = []
    mother_r: List = []
    matgen_r: List = []
    name_r: List = []
    sex_r: List = []
    pupgen_r: List = []
    age_r: List = []
    session_r: List = []
    rec_num_r: List = []
    audio_files: List[Path] = []
    total_calls = 0
    processed_wavs_since_gc = 0
    last_rate: int = 250000

    row_wall_times: deque = deque(maxlen=48)

    old_cwd = os.getcwd()
    resolve_wav_path_fn = __import__(
        "utils.audio_paths", fromlist=["resolve_wav_path"]
    ).resolve_wav_path

    welch_sanity_done = False

    def _ensure_welch_sanity(signal_arr, rate_hz: int) -> None:
        """Fail fast on the first loaded recording if Welch/PSD path cannot run."""
        nonlocal welch_sanity_done
        if welch_sanity_done:
            return
        from steps.preflight_welch import welch_sanity_check_signal  # type: ignore

        welch_sanity_check_signal(signal_arr, int(rate_hz))
        welch_sanity_done = True

    try:

        def estimate_row_eta(tr_total: int, tr_done: int) -> Optional[float]:
            """ETA from mean wall time per metadata row (includes skips and fast paths)."""
            if tr_total <= 0 or not row_wall_times:
                return None
            rem = max(0, tr_total - tr_done)
            if rem <= 0:
                return None
            return float(rem * (sum(row_wall_times) / len(row_wall_times)))

        if client_metadata_files:
            total_rows = 0
            for mf in client_metadata_files:
                try:
                    data = read_metadata_as_lists(str(mf))
                    total_rows += len(data["Mother"])
                except Exception:
                    continue
            rows_done = 0
            sel_resolved = selected.resolve()
            for mf in client_metadata_files:
                try:
                    data = read_metadata_as_lists(str(mf))
                except Exception:
                    summary.error_messages.append(f"Unreadable metadata: {mf.name}")
                    continue
                n = len(data["Mother"])
                for i in range(n):
                    row_t0 = time.perf_counter()
                    progress_msg: Optional[str] = None
                    try:
                        rows_done += 1
                        m, mg, n_, sx, pg, d, ses, rec = _meta_row(data, i)
                        summary.metadata_rows_scanned += 1
                        if mf.parent.resolve() == sel_resolved:
                            wav = resolve_wav_under_year_folder(
                                sel_resolved, m, mg, n_, pg, d, ses, rec
                            )
                        else:
                            wav = _resolve_wav_under_mother_folder(
                                mf.parent, m, mg, n_, pg, d, ses, rec, resolve_wav_path_fn
                            )
                        if wav is None:
                            sx_nf = _excel_sex_with_pup_fallback(sx, m, n_, sex_lookup)
                            inventory.append(
                                MetadataRow(
                                    year,
                                    mf.name,
                                    m,
                                    mg,
                                    n_,
                                    sx_nf,
                                    pg,
                                    d,
                                    ses,
                                    rec,
                                    "",
                                    "",
                                    "Recording file not found",
                                )
                            )
                            continue
                        wav = wav.resolve()
                        sx = _sex_for_metadata_row(sx, m, n_, sex_lookup, sel_resolved, wav)
                        display_name = _output_pup_name_for_segmentation(
                            sel_resolved, wav, n_
                        )
                        if _is_year_only_folder(selected):
                            try:
                                wav.relative_to(sel_resolved)
                            except ValueError:
                                continue

                        if not wav_matches_subfolder_prefixes(
                            wav, selected, dataset_root, year, subfolder_prefixes
                        ):
                            inventory.append(
                                MetadataRow(
                                    year,
                                    mf.name,
                                    m,
                                    mg,
                                    display_name,
                                    sx,
                                    pg,
                                    d,
                                    ses,
                                    rec,
                                    str(wav),
                                    _excel_path_column(selected, wav, year),
                                    "Skipped (subfolder filter)",
                                    0,
                                )
                            )
                            progress_msg = (
                                f"[{year}] Skip (subfolder filter): {_wav_log_path(wav)}"
                            )
                            continue

                        summary.wav_files_found += 1
                        try:
                            signal, rate = librosa.load(str(wav), sr=250000)
                            last_rate = int(rate)
                            _ensure_welch_sanity(signal, last_rate)
                            calls = segment_single_recording(
                                signal=signal,
                                Fs=rate,
                                frame_length=FRAME_LENGTH,
                                overlap=OVERLAP,
                                thresh=THRESH,
                                harmony_th=HARMONY_TH,
                                signal_file_name=str(wav),
                            )
                        except Exception as exc:
                            summary.wav_segmentation_failed += 1
                            summary.error_messages.append(f"{_wav_log_path(wav)}: {exc}")
                            inventory.append(
                                MetadataRow(
                                    year,
                                    mf.name,
                                    m,
                                    mg,
                                    display_name,
                                    sx,
                                    pg,
                                    d,
                                    ses,
                                    rec,
                                    str(wav),
                                    _excel_path_column(selected, wav, year),
                                    f"Error: {exc}",
                                    0,
                                )
                            )
                            continue

                        del signal
                        processed_wavs_since_gc += 1
                        if processed_wavs_since_gc >= 250:
                            gc.collect()
                            processed_wavs_since_gc = 0

                        summary.wav_segmentation_succeeded += 1
                        if not calls:
                            summary.recordings_with_zero_syllables += 1

                        audio_paths.append(wav)
                        mother_r.append(m)
                        matgen_r.append(mg)
                        name_r.append(display_name)
                        sex_r.append(sx)
                        pupgen_r.append(pg)
                        age_r.append(int(d))
                        session_r.append(int(ses))
                        rec_num_r.append(rec)
                        audio_files.append(wav)

                        path_excel = _excel_path_column(selected, wav, year)
                        inventory.append(
                            MetadataRow(
                                year,
                                mf.name,
                                m,
                                mg,
                                display_name,
                                sx,
                                pg,
                                d,
                                ses,
                                rec,
                                str(wav),
                                path_excel,
                                "OK",
                                len(calls),
                            )
                        )

                        for call in calls:
                            st, en = float(call[0]), float(call[1])
                            sheet.append(
                                [
                                    path_excel,
                                    m,
                                    mg,
                                    display_name,
                                    sx,
                                    pg,
                                    int(d),
                                    int(ses),
                                    rec,
                                    st,
                                    en,
                                    en - st,
                                ]
                            )
                            total_calls += 1

                        progress_msg = (
                            f"[{year}] Segmenting recording {rows_done}/{total_rows}: {_wav_log_path(wav)}"
                        )
                    finally:
                        row_wall_times.append(
                            min(300.0, max(0.0, time.perf_counter() - row_t0))
                        )
                    if progress_msg:
                        _emit_progress(
                            progress,
                            span_t(0.55 * (rows_done / max(1, total_rows))),
                            progress_msg,
                            None,
                        )

        elif dataset_root is not None and metadata_dir is not None:
            os.chdir(str(dataset_root))
            meta = _load_merged_metadata(metadata_dir, year)
            n_meta = len(meta.get("Mother", [])) if meta else 0
            rows_done = 0
            if n_meta > 0:
                from utils.audio_paths import resolve_wav_usv_recordings_layout  # type: ignore

                year_root_resolved = selected.resolve()
                for i in range(n_meta):
                    row_t0 = time.perf_counter()
                    progress_msg: Optional[str] = None
                    try:
                        rows_done += 1
                        m, mg, n_, sx, pg, d, ses, rec = _meta_row(meta, i)
                        summary.metadata_rows_scanned += 1
                        wav = resolve_wav_usv_recordings_layout(
                            "USV_Recordings",
                            year,
                            m,
                            mg,
                            n_,
                            pg,
                            int(d),
                            int(ses),
                            rec,
                        )
                        if wav is None:
                            sx_nf = _excel_sex_with_pup_fallback(sx, m, n_, sex_lookup)
                            inventory.append(
                                MetadataRow(
                                    year,
                                    "(metadata)",
                                    m,
                                    mg,
                                    n_,
                                    sx_nf,
                                    pg,
                                    d,
                                    ses,
                                    rec,
                                    "",
                                    "",
                                    "Recording file not found",
                                )
                            )
                            continue
                        wav = wav.resolve()
                        sx = _sex_for_metadata_row(
                            sx, m, n_, sex_lookup, year_root_resolved, wav
                        )
                        display_name = _output_pup_name_for_segmentation(
                            year_root_resolved, wav, n_
                        )
                        if _is_year_only_folder(selected):
                            try:
                                wav.relative_to(selected.resolve())
                            except ValueError:
                                continue

                        if not wav_matches_subfolder_prefixes(
                            wav, selected, dataset_root, year, subfolder_prefixes
                        ):
                            inventory.append(
                                MetadataRow(
                                    year,
                                    "(metadata)",
                                    m,
                                    mg,
                                    display_name,
                                    sx,
                                    pg,
                                    d,
                                    ses,
                                    rec,
                                    str(wav),
                                    _excel_path_from_resolved_wav(dataset_root, wav, year),
                                    "Skipped (subfolder filter)",
                                    0,
                                )
                            )
                            progress_msg = (
                                f"[{year}] Skip (subfolder filter): {_wav_log_path(wav)}"
                            )
                            continue

                        summary.wav_files_found += 1
                        try:
                            signal, rate = librosa.load(str(wav), sr=250000)
                            last_rate = int(rate)
                            _ensure_welch_sanity(signal, last_rate)
                            calls = segment_single_recording(
                                signal=signal,
                                Fs=rate,
                                frame_length=FRAME_LENGTH,
                                overlap=OVERLAP,
                                thresh=THRESH,
                                harmony_th=HARMONY_TH,
                                signal_file_name=str(wav),
                            )
                        except Exception as exc:
                            summary.wav_segmentation_failed += 1
                            summary.error_messages.append(f"{_wav_log_path(wav)}: {exc}")
                            inventory.append(
                                MetadataRow(
                                    year,
                                    "(metadata)",
                                    m,
                                    mg,
                                    display_name,
                                    sx,
                                    pg,
                                    d,
                                    ses,
                                    rec,
                                    str(wav),
                                    _excel_path_from_resolved_wav(dataset_root, wav, year),
                                    f"Error: {exc}",
                                    0,
                                )
                            )
                            continue

                        del signal
                        processed_wavs_since_gc += 1
                        if processed_wavs_since_gc >= 250:
                            gc.collect()
                            processed_wavs_since_gc = 0
                        summary.wav_segmentation_succeeded += 1
                        if not calls:
                            summary.recordings_with_zero_syllables += 1

                        audio_paths.append(wav)
                        mother_r.append(m)
                        matgen_r.append(mg)
                        name_r.append(display_name)
                        sex_r.append(sx)
                        pupgen_r.append(pg)
                        age_r.append(int(d))
                        session_r.append(int(ses))
                        rec_num_r.append(rec)
                        audio_files.append(wav)

                        path_excel = _excel_path_from_resolved_wav(dataset_root, wav, year)
                        inventory.append(
                            MetadataRow(
                                year,
                                "(metadata)",
                                m,
                                mg,
                                display_name,
                                sx,
                                pg,
                                d,
                                ses,
                                rec,
                                str(wav),
                                path_excel,
                                "OK",
                                len(calls),
                            )
                        )

                        for call in calls:
                            st, en = float(call[0]), float(call[1])
                            sheet.append(
                                [
                                    path_excel,
                                    m,
                                    mg,
                                    display_name,
                                    sx,
                                    pg,
                                    int(d),
                                    int(ses),
                                    rec,
                                    st,
                                    en,
                                    en - st,
                                ]
                            )
                            total_calls += 1

                        progress_msg = (
                            f"[{year}] Segmenting recording {rows_done}/{n_meta}: {_wav_log_path(wav)}"
                        )
                    finally:
                        row_wall_times.append(
                            min(300.0, max(0.0, time.perf_counter() - row_t0))
                        )
                    if progress_msg:
                        _emit_progress(
                            progress,
                            span_t(0.55 * (rows_done / max(1, n_meta))),
                            progress_msg,
                            None,
                        )

        if not audio_files:
            os.chdir(old_cwd)
            _emit_progress(
                progress,
                span_t(0.05),
                f"[{year}] Discovering audio files under folder…",
                None,
            )
            wav_list = _scan_audio_files(
                selected,
                progress=progress,
                progress_p=span_t(0.05),
                year=year,
                dataset_root=dataset_root,
                subfolder_prefixes=subfolder_prefixes,
            )
            if not wav_list:
                if not inventory and total_calls == 0:
                    raise ValueError(
                        f"[{year}] No audio recordings found and no usable metadata rows."
                    )
            else:
                filtered_wavs = [wp.resolve() for wp in wav_list]
                row_wall_times.clear()
                scan_n = max(1, len(filtered_wavs))
                for idx, wp in enumerate(filtered_wavs, start=1):
                    row_t0 = time.perf_counter()
                    progress_msg: Optional[str] = None
                    try:
                        summary.metadata_rows_scanned += 1
                        summary.wav_files_found += 1
                        try:
                            signal, rate = librosa.load(str(wp), sr=250000)
                            last_rate = int(rate)
                            _ensure_welch_sanity(signal, last_rate)
                            calls = segment_single_recording(
                                signal=signal,
                                Fs=rate,
                                frame_length=FRAME_LENGTH,
                                overlap=OVERLAP,
                                thresh=THRESH,
                                harmony_th=HARMONY_TH,
                                signal_file_name=str(wp),
                            )
                        except Exception as exc:
                            summary.wav_segmentation_failed += 1
                            summary.error_messages.append(f"{_wav_log_path(wp)}: {exc}")
                            continue

                        del signal
                        processed_wavs_since_gc += 1
                        if processed_wavs_since_gc >= 250:
                            gc.collect()
                            processed_wavs_since_gc = 0
                        summary.wav_segmentation_succeeded += 1
                        if not calls:
                            summary.recordings_with_zero_syllables += 1

                        m, mg, n_, sx_guess, pg, d, ses, rec = _extract_row_metadata_from_path_layout(
                            selected, wp
                        )
                        raw_pup = _raw_pup_folder_from_path(selected, wp)
                        path_name_hints = _pup_path_hints_for_wav(selected, wp)
                        details = _resolve_pup_details(m, raw_pup, n_, pup_details_lookup)
                        if details is not None:
                            if str(details.get("offspring_genotype", "")).strip():
                                pg = _normalize_genotype_token(str(details["offspring_genotype"]))
                            ds = str(details.get("sex", "")).strip()
                            if ds:
                                from utils.io_utils import normalize_sex_cell  # type: ignore

                                sx = normalize_sex_cell(ds)
                            else:
                                sx = _resolve_sex_from_pup_tables(
                                    m,
                                    n_,
                                    sex_lookup,
                                    extra_name_hints=path_name_hints,
                                )
                        else:
                            sx = _sex_for_metadata_row(
                                sx_guess, m, n_, sex_lookup, selected, wp
                            )
                        audio_paths.append(wp)
                        mother_r.append(m)
                        matgen_r.append(mg)
                        name_r.append(n_)
                        sex_r.append(sx)
                        pupgen_r.append(pg)
                        age_r.append(d)
                        session_r.append(ses)
                        rec_num_r.append(rec)
                        audio_files.append(wp)
                        path_excel = _excel_path_column(selected, wp, year)
                        inventory.append(
                            MetadataRow(
                                year,
                                "(scan)",
                                m,
                                mg,
                                n_,
                                sx,
                                pg,
                                d,
                                ses,
                                rec,
                                str(wp),
                                path_excel,
                                "OK",
                                len(calls),
                            )
                        )
                        for call in calls:
                            st, en = float(call[0]), float(call[1])
                            sheet.append(
                                [
                                    path_excel,
                                    m,
                                    mg,
                                    n_,
                                    sx,
                                    pg,
                                    d,
                                    ses,
                                    rec,
                                    st,
                                    en,
                                    en - st,
                                ]
                            )
                            total_calls += 1
                        progress_msg = (
                            f"[{year}] Segmenting recording {idx}/{scan_n}: {_wav_log_path(wp)}"
                        )
                    finally:
                        row_wall_times.append(
                            min(300.0, max(0.0, time.perf_counter() - row_t0))
                        )
                    if progress_msg:
                        _emit_progress(
                            progress,
                            span_t(0.55 * (idx / scan_n)),
                            progress_msg,
                            None,
                        )

        try:
            os.chdir(old_cwd)
        except OSError:
            pass

        summary.total_syllable_rows = total_calls

        if total_calls == 0 and not audio_files:
            hint = ""
            if subfolder_prefixes:
                hint = " Try clearing subfolder filters or check that metadata paths match the selected folders."
            raise ValueError(f"[{year}] No recordings processed.{hint}")

        ts = output_timestamp_suffix()
        out_stem = "segmentation_classification" if run_classification else "segmentation"
        output_path = outputs_dir / f"{out_stem}_{year}_{ts}.xlsx"
        outputs_dir.mkdir(parents=True, exist_ok=True)
        if run_classification:
            p_save, p_read, p_feat, p_cls_lo, p_pre_enrich, p_done = (
                0.58,
                0.62,
                0.70,
                0.82,
                0.94,
                1.0,
            )
        else:
            p_save, p_read, p_feat, p_cls_lo, p_pre_enrich, p_done = (
                0.58,
                0.62,
                0.74,
                0.82,
                0.88,
                1.0,
            )
        _emit_progress(progress, span_t(p_save), f"[{year}] Saving segmentation workbook...")
        book.save(str(output_path))
        out_str = str(output_path.resolve())
        siz = len(audio_paths)

        _emit_progress(progress, span_t(p_read), f"[{year}] Reading segmentation table...")
        (
            path_syl,
            mother_syl,
            matgen_syl,
            name_syl,
            sex_syl,
            pupgen_syl,
            age_syl,
            session_syl,
            rec_num_syl,
            start_syl,
            end_syl,
        ) = read_segmentation_results(out_str, logger=None)

        _emit_progress(progress, span_t(p_feat), f"[{year}] Computing ISI and frequencies...")
        compute_basic_features(
            file_path=out_str,
            signal_vec=None,
            siz=siz,
            mother=mother_r,
            name=name_r,
            age=age_r,
            session=session_r,
            rec_num=rec_num_r,
            mother_syl=mother_syl,
            name_syl=name_syl,
            age_syl=age_syl,
            session_syl=session_syl,
            rec_num_syl=rec_num_syl,
            start_syl=start_syl,
            end_syl=end_syl,
            rate=last_rate,
            logger=None,
            audio_paths=audio_paths,
        )

        if run_classification:
            cls_root = _classification_year_root(selected, year, dataset_root)
            model_path, model_tried_paths = _resolve_model_path()
            year_audio = cls_root if cls_root.is_dir() else None
            cnn_span = max(1e-6, p_pre_enrich - p_cls_lo)

            if model_path is not None:
                summary.classification_model_path = str(model_path.resolve())
                _emit_progress(
                    progress,
                    span_t(p_cls_lo),
                    f"[{year}] Syllable classification…",
                    None,
                )
                cnn_hook_phase: List[str] = ["resolve_paths"]

                def on_cnn_progress(
                    done: int,
                    total: int,
                    phase: str = "spectrograms",
                    chunk_end: Optional[int] = None,
                ) -> None:
                    phase_bucket = (
                        "predict"
                        if phase in ("predict", "predict_chunk")
                        else phase
                    )
                    if phase_bucket != cnn_hook_phase[0]:
                        cnn_hook_phase[0] = phase_bucket
                    frac_done = min(1.0, done / max(1, total))
                    if phase == "resolve_paths":
                        frac_in_cnn = 0.05 * frac_done
                        msg = f"[{year}] Classify: resolving paths {done}/{total}"
                    elif phase == "spectrograms":
                        frac_in_cnn = 0.05 + 0.50 * frac_done
                        msg = f"[{year}] Build spectrograms {done}/{total}"
                    elif phase == "predict_chunk":
                        ce = int(chunk_end) if chunk_end is not None else done
                        ce = min(ce, total)
                        lo = min(total, done + 1)
                        frac_in_cnn = 0.55 + 0.45 * (done / max(1, total))
                        hint = (
                            " — first batches can take many minutes (model/GPU warmup)"
                            if done == 0
                            else ""
                        )
                        msg = (
                            f"[{year}] Classify: inference in progress "
                            f"(syllables {lo}–{ce} of {total}){hint}…"
                        )
                    else:
                        frac_in_cnn = 0.55 + 0.45 * frac_done
                        msg = f"[{year}] Classify {done}/{total}"
                    p_inner = p_cls_lo + cnn_span * frac_in_cnn
                    _emit_progress(
                        progress,
                        span_t(p_inner),
                        msg,
                        None,
                    )

                def on_classify_stage(phase: str) -> None:
                    labels = {
                        "load_model": f"[{year}] Classify: loading model (TensorFlow)…",
                        "model_ready": f"[{year}] Classify: model ready",
                        "postprocess": f"[{year}] Finishing classification (post-processing labels)…",
                        "write_excel": f"[{year}] Writing syllable labels to Excel (large file, please wait)…",
                        "write_done": f"[{year}] Syllable labels saved.",
                    }
                    msg = labels.get(phase, phase)
                    frac = {
                        "load_model": 0.02,
                        "model_ready": 0.04,
                        "postprocess": 0.97,
                        "write_excel": 0.99,
                        "write_done": 1.0,
                    }.get(phase, 1.0)
                    inner_p = p_cls_lo + cnn_span * min(1.0, float(frac))
                    _emit_progress(progress, span_t(inner_p), msg, None)

                try:
                    run_syllable_classification(
                        file_path=out_str,
                        year=year,
                        model_path=str(model_path),
                        age_syl=age_syl,
                        matgen_syl=matgen_syl,
                        pupgen_syl=pupgen_syl,
                        mother_syl=mother_syl,
                        name_syl=name_syl,
                        sex_syl=sex_syl,
                        session_syl=session_syl,
                        rec_num_syl=rec_num_syl,
                        start_syl=start_syl,
                        end_syl=end_syl,
                        wav_path_syl=path_syl,
                        logger=None,
                        year_audio_root=year_audio,
                        progress_hook=on_cnn_progress,
                        save_npy=False,
                        stage_callback=on_classify_stage,
                    )
                except Exception as exc:
                    summary.error_messages.append(f"Classification: {exc}")
                    _write_constant_syllable_column(out_str, 10)
            else:
                tried_txt = "; ".join(model_tried_paths) if model_tried_paths else "(no candidates)"
                summary.error_messages.append(
                    "Classification skipped (model_weights.h6 not found). "
                    f"Set USV_MODEL_PATH or place the file under segmentation-app/models/. Checked: {tried_txt}"
                )
                _write_constant_syllable_column(out_str, 10)

        _emit_progress(
            progress,
            span_t(p_pre_enrich),
            f"[{year}] Enriching columns (large workbooks may take several minutes)…",
            None,
        )
        enrich_segmentation_columns(
            file_path=out_str,
            year=year,
            logger=None,
            include_syllable_classification_columns=run_classification,
        )
        mid_done = p_pre_enrich + (p_done - p_pre_enrich) * 0.88
        _emit_progress(
            progress,
            span_t(mid_done),
            f"[{year}] Enrichment saved; updating metadata counts…",
            None,
        )

        merge_syllable_counts_from_excel(inventory, out_str)

        summary.output_files.append(out_str)
        _emit_progress(
            progress,
            span_t(p_done),
            f"[{year}] Done. Syllables: {total_calls}. Output: {out_str}",
            None,
        )
        return out_str, summary, inventory

    finally:
        try:
            os.chdir(old_cwd)
        except OSError:
            pass


def run_pipeline(options: PipelineOptions, progress: ProgressFn) -> Tuple[RunSummary, List[str]]:
    """
    Main entry: multi-year aware, optional metadata-only, dual Excel outputs.
    """
    root = Path(options.root_folder).resolve()
    if not root.is_dir():
        raise ValueError(f"Selected folder does not exist: {root}")

    out_dir = Path(options.output_dir).resolve() if options.output_dir else Path.cwd() / "outputs"
    out_dir.mkdir(parents=True, exist_ok=True)

    last_emitted_bar_p = 0.0

    def _emit_overall_progress(
        p: float,
        msg: str,
        stage_eta_seconds: Optional[float] = None,
    ) -> None:
        """Emit overall pipeline progress (no ETA — UI shows elapsed time only)."""
        nonlocal last_emitted_bar_p
        _ = stage_eta_seconds  # unused; kept for call-site compatibility
        p_clamped = max(0.0, min(1.0, float(p)))
        if p_clamped + 1e-9 < last_emitted_bar_p:
            p_clamped = last_emitted_bar_p
        last_emitted_bar_p = p_clamped
        _emit_progress(progress, p_clamped, msg, None)

    _emit_overall_progress(0.0, "Initializing preprocessing modules...")
    preprocessing_dir = _runtime_base_dir() / "preprocessing"
    if not preprocessing_dir.exists():
        raise ValueError("Missing `preprocessing` directory next to app files.")
    _add_preprocessing_to_path(preprocessing_dir)

    pairs = discover_year_roots(root)
    if options.years:
        allow = set(options.years)
        pairs = [(y, p) for y, p in pairs if y in allow]
        if not pairs:
            raise ValueError("No matching years after filter.")

    total_summary = RunSummary()
    all_outputs: List[str] = []
    n_years = len(pairs)
    per_year_summary: List[Dict[str, Any]] = []

    def _progress_year_eta_global_only(p: float, msg: str, eta: Optional[float] = None) -> None:
        """Row/phase ETAs from process_single_year are for the current year only; blend distorts multi-year ETA."""
        if n_years > 1:
            eta = None
        _emit_overall_progress(p, msg, eta)

    years_with_outputs = 0
    years_skipped_by_empty_tree = 0
    for yi, (year_str, year_path) in enumerate(pairs):
        lo = yi / n_years
        hi = (yi + 1) / n_years
        subfolder_prefixes = _opt_year_subfolders(options.subfolder_filters, year_str)
        explicit_meta_for_year: Optional[Path] = None
        if options.metadata_file_overrides and year_str in options.metadata_file_overrides:
            p = str(options.metadata_file_overrides[year_str]).strip()
            if p:
                explicit_meta_for_year = Path(p)
        _emit_overall_progress(lo, f"Year {year_str}: starting...")
        if subfolder_prefixes is not None and len(subfolder_prefixes) == 0:
            _emit_overall_progress(
                hi,
                f"Year {year_str}: no folders selected in the tree — skipped.",
            )
            years_skipped_by_empty_tree += 1
            continue

        total_summary.years_processed.append(year_str)

        try:
            if options.metadata_only:
                from utils.audio_paths import resolve_wav_path  # type: ignore

                inv_rows, part = collect_metadata_inventory_only(
                    selected=year_path,
                    year=year_str,
                    resolve_wav_path_fn=resolve_wav_path,
                    progress=progress,
                    progress_base=lo,
                    progress_span=hi - lo,
                    subfolder_prefixes=subfolder_prefixes,
                    explicit_metadata_file=explicit_meta_for_year,
                )
                meta_path = str(out_dir / f"recordings_metadata_{year_str}_{output_timestamp_suffix()}.xlsx")
                _emit_overall_progress(lo + (hi - lo) * 0.95, f"Year {year_str}: writing metadata workbook...")
                save_metadata_inventory(meta_path, inv_rows)
                total_summary.merge(part)
                total_summary.output_files.append(meta_path)
                all_outputs.append(meta_path)
                per_year_summary.append(_inventory_year_summary_row(inv_rows, year_str))
                years_with_outputs += 1
                _emit_overall_progress(hi, f"Year {year_str} metadata inventory saved.")
                continue

            syllable_path: Optional[str] = None
            inv_rows: List[MetadataRow] = []
            part = RunSummary()

            if options.want_syllables_xlsx:
                syllable_path, part, inv_rows = process_single_year(
                    selected=year_path,
                    year=year_str,
                    outputs_dir=out_dir,
                    progress=_progress_year_eta_global_only,
                    progress_lo=lo,
                    progress_hi=hi,
                    subfolder_prefixes=subfolder_prefixes,
                    explicit_metadata_file=explicit_meta_for_year,
                    run_classification=options.run_classification,
                )
                total_summary.merge(part)
                if syllable_path:
                    all_outputs.append(syllable_path)
            else:
                from utils.audio_paths import resolve_wav_path  # type: ignore

                inv_rows, part = collect_metadata_inventory_only(
                    selected=year_path,
                    year=year_str,
                    resolve_wav_path_fn=resolve_wav_path,
                    progress=_emit_overall_progress,
                    progress_base=lo,
                    progress_span=(hi - lo) * 0.9,
                    subfolder_prefixes=subfolder_prefixes,
                    explicit_metadata_file=explicit_meta_for_year,
                )
                total_summary.merge(part)

            if options.want_metadata_xlsx:
                if not inv_rows and syllable_path is None:
                    from utils.audio_paths import resolve_wav_path  # type: ignore

                    inv_rows, part = collect_metadata_inventory_only(
                        selected=year_path,
                        year=year_str,
                        resolve_wav_path_fn=resolve_wav_path,
                    progress=_emit_overall_progress,
                        progress_base=lo + (hi - lo) * 0.5,
                        progress_span=(hi - lo) * 0.5,
                        subfolder_prefixes=subfolder_prefixes,
                        explicit_metadata_file=explicit_meta_for_year,
                    )
                    total_summary.merge(part)
                if syllable_path and inv_rows:
                    _emit_overall_progress(lo + (hi - lo) * 0.96, f"Year {year_str}: merging syllable counts...")
                    merge_syllable_counts_from_excel(inv_rows, syllable_path)
                meta_path = str(out_dir / f"recordings_metadata_{year_str}_{output_timestamp_suffix()}.xlsx")
                _emit_overall_progress(lo + (hi - lo) * 0.98, f"Year {year_str}: writing metadata workbook...")
                save_metadata_inventory(meta_path, inv_rows)
                total_summary.output_files.append(meta_path)
                all_outputs.append(meta_path)
            per_year_summary.append(_inventory_year_summary_row(inv_rows, year_str))
            years_with_outputs += 1
        except Exception as exc:
            msg = f"[{year_str}] {exc}"
            total_summary.error_messages.append(msg)
            _emit_overall_progress(hi, f"Year {year_str}: failed ({exc}) — continuing to next year.")
            continue

    seg_merge_sources = _ordered_unique_per_year_segmentation_paths(total_summary.output_files)
    if len(seg_merge_sources) > 1:
        merge_stem = (
            "segmentation_classification"
            if options.run_classification
            else "segmentation"
        )
        merged_path = out_dir / f"{merge_stem}_Multiple_Years_{output_timestamp_suffix()}.xlsx"
        _merge_segmentation_workbooks(seg_merge_sources, merged_path)
        merged_s = str(merged_path.resolve())
        seg_set = set(seg_merge_sources)
        total_summary.output_files = [p for p in total_summary.output_files if p not in seg_set]
        total_summary.output_files.append(merged_s)
        all_outputs = [p for p in all_outputs if p not in seg_set]
        all_outputs.append(merged_s)
        for p in seg_merge_sources:
            try:
                Path(p).unlink(missing_ok=True)
            except OSError:
                pass

    if per_year_summary:
        base_wb = _segmentation_workbook_for_summary_name(total_summary.output_files)
        if base_wb is None:
            base_wb = _metadata_workbook_for_summary_name(total_summary.output_files)
        if base_wb is not None:
            summary_path = base_wb.parent / f"{base_wb.stem}_summary.xlsx"
        else:
            summary_path = out_dir / f"processing_summary_{output_timestamp_suffix()}.xlsx"
        save_processing_summary_workbook(str(summary_path), per_year_summary)
        summary_abs = str(summary_path.resolve())
        total_summary.output_files.append(summary_abs)
        all_outputs.append(summary_abs)

    total_summary.output_directory = str(out_dir.resolve())
    if years_with_outputs == 0:
        if years_skipped_by_empty_tree == len(pairs):
            raise ValueError(
                "No output files were produced: all selected years were skipped "
                "because no folders are checked in the tree."
            )
        notes = "; ".join(total_summary.error_messages[:5]) if total_summary.error_messages else ""
        if notes:
            raise ValueError(
                "No output files were produced for the selected years/folders. "
                f"First errors: {notes}"
            )
        raise ValueError("No output files were produced for the selected years/folders.")
    _emit_overall_progress(1.0, "All selected years finished.", None)
    return total_summary, all_outputs


def execute_pipeline(
    folder_path: str,
    progress_callback: ProgressFn,
    *,
    output_dir: Optional[str] = None,
    years: Optional[List[str]] = None,
    want_syllables_xlsx: bool = True,
    want_metadata_xlsx: bool = True,
    metadata_only: bool = False,
    run_classification: bool = True,
    subfolder_filters: Optional[Dict[str, List[str]]] = None,
    metadata_file_overrides: Optional[Dict[str, str]] = None,
) -> Tuple[str, RunSummary]:
    """
    Backward-compatible wrapper. Returns (primary syllable xlsx path or first output, summary).

    ``progress_callback`` may be ``(progress, message)`` or ``(progress, message, eta_seconds)``.
    """
    opts = PipelineOptions(
        root_folder=folder_path,
        output_dir=output_dir,
        years=years,
        want_syllables_xlsx=want_syllables_xlsx,
        want_metadata_xlsx=want_metadata_xlsx,
        metadata_only=metadata_only,
        run_classification=run_classification,
        subfolder_filters=subfolder_filters,
        metadata_file_overrides=metadata_file_overrides,
    )

    def _wrap(p: float, msg: str, eta: Optional[float] = None) -> None:
        try:
            progress_callback(p, msg, eta)
        except TypeError:
            progress_callback(p, msg)

    summary, outputs = run_pipeline(opts, _wrap)
    primary = ""
    for p in outputs:
        if _is_segmentation_workbook_path(p):
            primary = p
            break
    if not primary and outputs:
        primary = outputs[0]
    return primary, summary
